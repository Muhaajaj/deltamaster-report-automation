#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
DeltaMaster Report Merge (TopM + Addison)
- Reads two Excel exports
- Builds cost-center level KPIs + “Modifikationen” logic
- Merges with Addison figures (Umsatzerlöse/Aufwendungen/Rohergebnis)
- Exports a formatted Excel report with highlighted columns

Data privacy:
- Do not commit real employer exports.
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import List

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# -----------------------------
# Helpers
# -----------------------------
def require_columns(df: pd.DataFrame, required: List[str], df_name: str) -> None:
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(
            f"{df_name} is missing required columns: {missing}\n"
            f"Available columns: {list(df.columns)}"
        )


def safe_div(numerator: pd.Series, denominator: pd.Series) -> pd.Series:
    denom = denominator.replace({0: np.nan})
    return numerator / denom


# -----------------------------
# Load + transform TopM export
# -----------------------------
def load_topm_report(path: Path) -> pd.DataFrame:
    # Uses last sheet, skips first 6 rows of meta/header like your Colab version
    df = pd.read_excel(path, sheet_name=-1, header=6)
    if df.shape[1] < 2:
        raise ValueError("TopM report looks empty or has unexpected structure.")
    df = df.copy()
    df.rename(columns={df.columns[0]: "Hilfsmittel", df.columns[1]: "Filiale"}, inplace=True)
    df["KSt"] = df["Filiale"].astype(str).str[:5]
    return df


def transform_topm(df: pd.DataFrame) -> pd.DataFrame:
    # Filter out “Sammelpositionen”
    df_clean = df[~df["Hilfsmittel"].isin(["Alle Hilfsmittel", "08 - Einlagen"])].copy()

    # Required numeric columns (based on your logic)
    required_cols = [
        "(1) Umsatz-\nberechnung",
        "(6) DB I =\n(1) - (5)",
        "AP DB I mit FP",
    ]
    require_columns(df_clean, required_cols, "TopM report")

    # DBI % and AP DBI % (row level)
    df_clean["DBI_pct"] = safe_div(df_clean["(6) DB I =\n(1) - (5)"], df_clean["(1) Umsatz-\nberechnung"])
    df_clean["AP_DBI_pct_mit_FP"] = safe_div(df_clean["AP DB I mit FP"], df_clean["(1) Umsatz-\nberechnung"])

    # Modifikationen logic (10 & 18)
    df_clean["Modifikationen"] = np.where(
        df_clean["Hilfsmittel"].isin(["10 - Gehhilfen", "18 - Kranken-/ Behindertenfahrzeuge"]),
        df_clean["AP DB I mit FP"],
        df_clean["(6) DB I =\n(1) - (5)"]
    )
    df_clean["DBI_pct_Modifikationen"] = safe_div(df_clean["Modifikationen"], df_clean["(1) Umsatz-\nberechnung"])

    # Modifikationen 09 & 32 logic
    df_clean["Modifikationen_09_32"] = df_clean["Modifikationen"]
    mask_0932 = df_clean["Hilfsmittel"].isin([
        "09 - Elektrostimulationsgeräte",
        "32 - Therapeutische Bewegungsgeräte"
    ])
    df_clean.loc[mask_0932, "Modifikationen_09_32"] = (
        df_clean.loc[mask_0932, "(1) Umsatz-\nberechnung"] * 0.80
    )
    df_clean["DBI_pct_Modifikationen_09_32"] = safe_div(df_clean["Modifikationen_09_32"], df_clean["(1) Umsatz-\nberechnung"])

    return df_clean


def aggregate_topm_to_kst(df_clean: pd.DataFrame) -> pd.DataFrame:
    # Percent columns should NOT be summed
    pct_cols = ["DBI_pct", "AP_DBI_pct_mit_FP", "DBI_pct_Modifikationen", "DBI_pct_Modifikationen_09_32"]
    numeric_cols = df_clean.select_dtypes(include="number").columns.tolist()
    sum_cols = [c for c in numeric_cols if c not in pct_cols]

    df_kst = df_clean.groupby("KSt", as_index=False)[sum_cols].sum()

    # Recompute pct on aggregated sums
    df_kst["DBI_pct"] = safe_div(df_kst["(6) DB I =\n(1) - (5)"], df_kst["(1) Umsatz-\nberechnung"])
    df_kst["AP_DBI_pct_mit_FP"] = safe_div(df_kst["AP DB I mit FP"], df_kst["(1) Umsatz-\nberechnung"])
    df_kst["DBI_pct_Modifikationen"] = safe_div(df_kst["Modifikationen"], df_kst["(1) Umsatz-\nberechnung"])
    df_kst["DBI_pct_Modifikationen_09_32"] = safe_div(df_kst["Modifikationen_09_32"], df_kst["(1) Umsatz-\nberechnung"])

    # Representative Filiale per KSt (mode)
    filiale_map = df_clean.groupby("KSt")["Filiale"].agg(lambda x: x.mode().iloc[0])
    df_kst["Filiale"] = df_kst["KSt"].map(filiale_map)

    return df_kst


# -----------------------------
# Load + transform Addison export
# -----------------------------
def load_addison_report(path: Path) -> pd.DataFrame:
    # Many DeltaMaster exports have extra header lines; you used skiprows=8
    df2 = pd.read_excel(path, sheet_name=-1, skiprows=8)
    return df2


def transform_addison(df2: pd.DataFrame) -> pd.DataFrame:
    # Expect these columns to exist after import
    required = ["KSt", "Art", "Wert4", "Wert6"]
    require_columns(df2, required, "Addison report")

    relevant_arten = ["Umsatzerlöse", "Aufwendungen für bez. Lfg. und Lst.", "Rohergebnis"]
    df_sub = df2[df2["Art"].isin(relevant_arten)].copy()

    pivot4 = df_sub.pivot_table(index="KSt", columns="Art", values="Wert4", aggfunc="sum")
    pivot6 = df_sub.pivot_table(index="KSt", columns="Art", values="Wert6", aggfunc="sum")

    df_new = pivot4.copy()
    df_new["Umsatzerlöse Kum"] = pivot6.get("Umsatzerlöse")
    df_new["Aufwendungen für bez. Lfg. und Lst. Kum"] = pivot6.get("Aufwendungen für bez. Lfg. und Lst.")
    df_new.reset_index(inplace=True)

    return df_new


# -----------------------------
# Merge + final calculations
# -----------------------------
def merge_and_calculate(df_topm_kst: pd.DataFrame, df_addison: pd.DataFrame) -> pd.DataFrame:
    df = df_topm_kst.merge(df_addison, on="KSt", how="left")

    # Fill missing values for calculation safety
    if "Umsatzerlöse" not in df.columns:
        df["Umsatzerlöse"] = 0.0
    if "Aufwendungen für bez. Lfg. und Lst." not in df.columns:
        df["Aufwendungen für bez. Lfg. und Lst."] = 0.0

    df["Umsatzerlöse"] = df["Umsatzerlöse"].fillna(0)
    df["Aufwendungen für bez. Lfg. und Lst."] = df["Aufwendungen für bez. Lfg. und Lst."].fillna(0)

    # Aufwendungen final = Umsatzerlöse * (1 - DBI_pct_Modifikationen) + Aufwendungen...
    df["Aufwendungen final"] = (
        df["Umsatzerlöse"] * (1 - df["DBI_pct_Modifikationen"]) +
        df["Aufwendungen für bez. Lfg. und Lst."]
    ).round(0)

    return df


def export_excel(df: pd.DataFrame, out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # Column order (keep your original intent but use clearer internal pct names)
    final_cols = [
        "KSt", "Filiale",
        "Aufträge",
        "(1) Umsatz-\nberechnung",
        "(2) Netto EK",
        "(3) Netto EK\nOhne WK",
        "(4) WK EK",
        "AP_EK_Verrechnung_WK_mit_FP",
        "(5) =\n(3) + (4)",
        "(6) DB I =\n(1) - (5)",
        "AP DB I mit FP",
        "Modifikationen",
        "Modifikationen_09_32",
        "DBI_pct",
        "AP_DBI_pct_mit_FP",
        "DBI_pct_Modifikationen",
        "DBI_pct_Modifikationen_09_32",
        "Umsatzerlöse",
        "Aufwendungen für bez. Lfg. und Lst.",
        "Rohergebnis",
        "Umsatzerlöse Kum",
        "Aufwendungen für bez. Lfg. und Lst. Kum",
        "Aufwendungen final"
    ]
    cols_existing = [c for c in final_cols if c in df.columns]
    df_export = df[cols_existing].copy()

    # Export first
    df_export.to_excel(out_path, index=False, sheet_name="Auswertung")

    # Highlight key columns
    wb = load_workbook(out_path)
    ws = wb["Auswertung"]
    header = [cell.value for cell in ws[1]]

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    columns_to_color = ["Aufwendungen final", "DBI_pct_Modifikationen"]

    for col_name in columns_to_color:
        if col_name in header:
            col_idx = header.index(col_name) + 1
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.fill = yellow_fill

    wb.save(out_path)


# -----------------------------
# CLI
# -----------------------------
def main() -> None:
    parser = argparse.ArgumentParser(description="Merge DeltaMaster TopM + Addison exports into one Excel report.")
    parser.add_argument("--topm", required=True, help="Path to TopM Excel export")
    parser.add_argument("--addison", required=True, help="Path to Addison Excel export")
    parser.add_argument("--out", default="outputs/Ergebnis_final_strukturiert.xlsx", help="Output Excel path")

    args = parser.parse_args()

    topm_path = Path(args.topm)
    addison_path = Path(args.addison)
    out_path = Path(args.out)

    df_topm = load_topm_report(topm_path)
    df_topm_clean = transform_topm(df_topm)
    df_topm_kst = aggregate_topm_to_kst(df_topm_clean)

    df_addison_raw = load_addison_report(addison_path)
    df_addison = transform_addison(df_addison_raw)

    df_final = merge_and_calculate(df_topm_kst, df_addison)
    export_excel(df_final, out_path)

    print(f"Done. Output written to: {out_path}")


if __name__ == "__main__":
    main()
